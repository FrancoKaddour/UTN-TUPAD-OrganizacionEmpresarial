"""
ChatBot WhatsApp - Gestion de Proveedores
La Basica Pasteleria - TPI Organizacion Empresarial
Stack: Python 3 + Flask + Twilio + openpyxl
"""

from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
import openpyxl
from openpyxl import load_workbook
from datetime import date
import os

app = Flask(__name__)

# ── Config ───────────────────────────────────────────────────────────────────
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "proveedores.xlsx")
SHEET_NAME = "Proveedores"
ROW_DATA = 6   # los datos arrancan en fila 6 (titulo en 2, subtitulo en 3, encabezados en 5)

CATEGORIAS = {
    "1": "Panaderia",
    "2": "Reposteria",
    "3": "Bebidas",
    "4": "Fiambres y Sandwicheria",
    "5": "Insumos de Cocina",
}

COLUMNAS = [
    "ID", "Nombre", "Categoria", "Productos",
    "Telefono", "Dias de Entrega", "Forma de Pago",
    "Estado", "Fecha Alta", "Ultima Modificacion",
]

# ── Estado en memoria (por numero de WhatsApp) ───────────────────────────────
# Cada entrada: { "paso": str, "datos": dict }
sesiones: dict[str, dict] = {}

# ── Helpers Excel ─────────────────────────────────────────────────────────────

def cargar_hoja():
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    return wb, ws


def siguiente_id(ws) -> str:
    """Genera el proximo ID correlativo tipo P007."""
    ids = []
    for row in ws.iter_rows(min_row=ROW_DATA, min_col=2, values_only=True):
        if row[0]:
            try:
                ids.append(int(str(row[0]).replace("P", "")))
            except ValueError:
                pass
    return f"P{(max(ids) + 1):03d}" if ids else "P001"


def buscar_proveedor(ws, identificador: str):
    """Busca por ID o por nombre (parcial, sin importar mayusculas)."""
    ident = identificador.strip().upper()
    for row in ws.iter_rows(min_row=ROW_DATA, min_col=2, values_only=True):
        if not row[0]:
            continue
        if str(row[0]).upper() == ident:
            return dict(zip(COLUMNAS, row))
        if ident in str(row[1]).upper():
            return dict(zip(COLUMNAS, row))
    return None


def actualizar_fila(ws, id_prov: str, campo: str, valor):
    col_idx = COLUMNAS.index(campo) + 2   # +2 porque datos arrancan en col B (indice 2)
    mod_idx = COLUMNAS.index("Ultima Modificacion") + 2
    for row in ws.iter_rows(min_row=ROW_DATA):
        if row[1].value and str(row[1].value).upper() == id_prov.upper():
            row[col_idx - 1].value = valor
            row[mod_idx - 1].value = date.today().isoformat()
            return True
    return False


def listar_proveedores(ws, categoria: str = None) -> list[dict]:
    resultado = []
    for row in ws.iter_rows(min_row=ROW_DATA, min_col=2, values_only=True):
        if not row[0]:
            continue
        prov = dict(zip(COLUMNAS, row))
        if prov["Estado"] != "Activo":
            continue
        if categoria and prov["Categoria"] != categoria:
            continue
        resultado.append(prov)
    return resultado


def formatear_proveedor(p: dict) -> str:
    return (
        f"*{p['ID']}* - {p['Nombre']}\n"
        f"  Categoria: {p['Categoria']}\n"
        f"  Productos: {p['Productos']}\n"
        f"  Telefono: {p['Telefono']}\n"
        f"  Entrega: {p['Dias de Entrega']}\n"
        f"  Pago: {p['Forma de Pago']}"
    )

# ── Menus ─────────────────────────────────────────────────────────────────────

MENU_PRINCIPAL = (
    "Hola! Soy el bot de proveedores de *La Basica Pasteleria*.\n\n"
    "Que deseas hacer?\n"
    "  *1* - Ver proveedores\n"
    "  *2* - Dar de alta un proveedor\n"
    "  *3* - Actualizar un proveedor\n"
    "  *4* - Dar de baja un proveedor\n"
    "  *0* - Salir"
)

MENU_CATEGORIAS = (
    "Elige una categoria:\n"
    "  *1* - Panaderia\n"
    "  *2* - Reposteria\n"
    "  *3* - Bebidas\n"
    "  *4* - Fiambres y Sandwicheria\n"
    "  *5* - Insumos de Cocina\n"
    "  *0* - Ver todos"
)

CAMPOS_ACTUALIZABLES = (
    "Que campo deseas actualizar?\n"
    "  *1* - Nombre\n"
    "  *2* - Categoria\n"
    "  *3* - Productos\n"
    "  *4* - Telefono\n"
    "  *5* - Dias de entrega\n"
    "  *6* - Forma de pago"
)

MAPA_CAMPOS = {
    "1": "Nombre",
    "2": "Categoria",
    "3": "Productos",
    "4": "Telefono",
    "5": "Dias de Entrega",
    "6": "Forma de Pago",
}

# ── Logica de conversacion ────────────────────────────────────────────────────

def procesar_mensaje(numero: str, texto: str) -> str:
    texto = texto.strip()
    sesion = sesiones.get(numero, {"paso": "inicio", "datos": {}})

    # ── Salida global ──────────────────────────────────────────────────────────
    if texto == "0" and sesion["paso"] not in ("menu",):
        sesiones[numero] = {"paso": "menu", "datos": {}}
        return MENU_PRINCIPAL

    # ── Maquina de estados ─────────────────────────────────────────────────────
    paso = sesion["paso"]
    datos = sesion["datos"]

    # ─ INICIO / MENU ──────────────────────────────────────────────────────────
    if paso in ("inicio", "menu"):
        if texto in ("hola", "hi", "inicio", "menu", "1", "2", "3", "4", "0"):
            if texto in ("hola", "hi", "inicio", "menu"):
                sesiones[numero] = {"paso": "menu", "datos": {}}
                return MENU_PRINCIPAL
            if texto == "0":
                sesiones.pop(numero, None)
                return "Hasta luego! Escribi *hola* para volver al menu."
            if texto == "1":
                sesiones[numero] = {"paso": "ver_filtro", "datos": {}}
                return MENU_CATEGORIAS
            if texto == "2":
                sesiones[numero] = {"paso": "alta_nombre", "datos": {}}
                return "Alta de proveedor.\nIngresa el *nombre* del proveedor:"
            if texto == "3":
                sesiones[numero] = {"paso": "actualizar_buscar", "datos": {}}
                return "Actualizacion de proveedor.\nIngresa el *ID o nombre* del proveedor:"
            if texto == "4":
                sesiones[numero] = {"paso": "baja_buscar", "datos": {}}
                return "Baja de proveedor.\nIngresa el *ID o nombre* del proveedor:"
        else:
            return f"Opcion no valida. {MENU_PRINCIPAL}"

    # ─ VER PROVEEDORES ────────────────────────────────────────────────────────
    if paso == "ver_filtro":
        wb, ws = cargar_hoja()
        if texto == "0":
            provs = listar_proveedores(ws)
            cat_txt = "Todos los proveedores activos"
        elif texto in CATEGORIAS:
            cat = CATEGORIAS[texto]
            provs = listar_proveedores(ws, cat)
            cat_txt = cat
        else:
            return f"Opcion invalida. {MENU_CATEGORIAS}"

        sesiones[numero] = {"paso": "menu", "datos": {}}

        if not provs:
            return f"No hay proveedores activos en '{cat_txt}'.\n\n{MENU_PRINCIPAL}"

        lineas = [f"*{cat_txt}* ({len(provs)} encontrados):\n"]
        for p in provs:
            lineas.append(formatear_proveedor(p))
            lineas.append("---")
        lineas.append(f"\nEscribi *menu* para volver al inicio.")
        return "\n".join(lineas)

    # ─ ALTA ───────────────────────────────────────────────────────────────────
    if paso == "alta_nombre":
        if len(texto) < 3:
            return "El nombre es muy corto. Ingresa el nombre del proveedor:"
        datos["Nombre"] = texto
        sesiones[numero] = {"paso": "alta_categoria", "datos": datos}
        return f"Categoria del proveedor '{texto}'?\n{MENU_CATEGORIAS.replace('*0* - Ver todos', '').strip()}"

    if paso == "alta_categoria":
        if texto not in CATEGORIAS:
            return f"Opcion invalida. Elige entre 1 y 5:\n{MENU_CATEGORIAS}"
        datos["Categoria"] = CATEGORIAS[texto]
        sesiones[numero] = {"paso": "alta_productos", "datos": datos}
        return "Productos que provee (separados por coma):"

    if paso == "alta_productos":
        if len(texto) < 3:
            return "Ingresa al menos un producto:"
        datos["Productos"] = texto
        sesiones[numero] = {"paso": "alta_telefono", "datos": datos}
        return "Numero de telefono del proveedor:"

    if paso == "alta_telefono":
        tel = texto.replace("-", "").replace(" ", "")
        if not tel.isdigit() or len(tel) < 7:
            return "Telefono invalido. Ingresa solo numeros (ej: 1154321001):"
        datos["Telefono"] = tel
        sesiones[numero] = {"paso": "alta_dias", "datos": datos}
        return "Dias de entrega (ej: Lunes, Miercoles, Viernes):"

    if paso == "alta_dias":
        if len(texto) < 3:
            return "Ingresa al menos un dia de entrega:"
        datos["Dias de Entrega"] = texto
        sesiones[numero] = {"paso": "alta_pago", "datos": datos}
        return "Forma de pago (Efectivo / Transferencia / Credito 30 dias):"

    if paso == "alta_pago":
        if len(texto) < 3:
            return "Ingresa la forma de pago:"
        datos["Forma de Pago"] = texto
        sesiones[numero] = {"paso": "alta_confirmar", "datos": datos}
        resumen = (
            f"Confirmas dar de alta este proveedor?\n\n"
            f"  Nombre: {datos['Nombre']}\n"
            f"  Categoria: {datos['Categoria']}\n"
            f"  Productos: {datos['Productos']}\n"
            f"  Telefono: {datos['Telefono']}\n"
            f"  Dias entrega: {datos['Dias de Entrega']}\n"
            f"  Forma pago: {datos['Forma de Pago']}\n\n"
            f"*SI* para confirmar / *NO* para cancelar"
        )
        return resumen

    if paso == "alta_confirmar":
        if texto.upper() == "SI":
            wb, ws = cargar_hoja()
            # G1: verificar duplicado por nombre
            if buscar_proveedor(ws, datos["Nombre"]):
                sesiones[numero] = {"paso": "menu", "datos": {}}
                return (
                    f"El proveedor '{datos['Nombre']}' ya existe en el sistema.\n"
                    f"Usa la opcion *3* para actualizar sus datos.\n\n{MENU_PRINCIPAL}"
                )
            nuevo_id = siguiente_id(ws)
            hoy = date.today().isoformat()
            # ws.append escribe desde columna A; como los datos estan en B,
            # anteponemos None para que la columna A quede vacia
            fila = [
                None,
                nuevo_id, datos["Nombre"], datos["Categoria"], datos["Productos"],
                datos["Telefono"], datos["Dias de Entrega"], datos["Forma de Pago"],
                "Activo", hoy, hoy,
            ]
            ws.append(fila)
            wb.save(EXCEL_FILE)
            sesiones[numero] = {"paso": "menu", "datos": {}}
            return (
                f"Proveedor dado de alta exitosamente!\n"
                f"ID asignado: *{nuevo_id}*\n\n{MENU_PRINCIPAL}"
            )
        elif texto.upper() == "NO":
            sesiones[numero] = {"paso": "menu", "datos": {}}
            return f"Alta cancelada.\n\n{MENU_PRINCIPAL}"
        else:
            return "Responde *SI* para confirmar o *NO* para cancelar:"

    # ─ ACTUALIZAR ─────────────────────────────────────────────────────────────
    if paso == "actualizar_buscar":
        wb, ws = cargar_hoja()
        prov = buscar_proveedor(ws, texto)
        if not prov:
            return f"No se encontro '{texto}'. Intenta con el ID (ej: P001) o el nombre exacto:"
        if prov["Estado"] == "Inactivo":
            sesiones[numero] = {"paso": "menu", "datos": {}}
            return f"El proveedor '{prov['Nombre']}' esta dado de baja y no puede modificarse.\n\n{MENU_PRINCIPAL}"
        datos["proveedor"] = prov
        sesiones[numero] = {"paso": "actualizar_campo", "datos": datos}
        return (
            f"Proveedor encontrado: *{prov['Nombre']}* ({prov['ID']})\n\n"
            f"{CAMPOS_ACTUALIZABLES}"
        )

    if paso == "actualizar_campo":
        if texto not in MAPA_CAMPOS:
            return f"Opcion invalida.\n{CAMPOS_ACTUALIZABLES}"
        datos["campo"] = MAPA_CAMPOS[texto]
        sesiones[numero] = {"paso": "actualizar_valor", "datos": datos}
        campo_display = datos["campo"].replace("_", " ")
        valor_actual = datos["proveedor"].get(datos["campo"], "-")
        if datos["campo"] == "Categoria":
            return f"Categoria actual: *{valor_actual}*\nElige nueva categoria:\n{MENU_CATEGORIAS}"
        return f"Valor actual de *{campo_display}*: {valor_actual}\nIngresa el nuevo valor:"

    if paso == "actualizar_valor":
        campo = datos["campo"]
        if campo == "Categoria":
            if texto in CATEGORIAS:
                nuevo_valor = CATEGORIAS[texto]
            else:
                return f"Opcion invalida. Elige entre 1 y 5:\n{MENU_CATEGORIAS}"
        elif campo == "Telefono":
            tel = texto.replace("-", "").replace(" ", "")
            if not tel.isdigit() or len(tel) < 7:
                return "Telefono invalido. Ingresa solo numeros:"
            nuevo_valor = tel
        else:
            if len(texto) < 2:
                return "Valor demasiado corto. Ingresa el nuevo valor:"
            nuevo_valor = texto

        datos["nuevo_valor"] = nuevo_valor
        campo_display = campo.replace("_", " ")
        sesiones[numero] = {"paso": "actualizar_confirmar", "datos": datos}
        return (
            f"Confirmas cambiar *{campo_display}* de '{datos['proveedor'][campo]}' a '{nuevo_valor}'?\n"
            f"*SI* para confirmar / *NO* para cancelar"
        )

    if paso == "actualizar_confirmar":
        if texto.upper() == "SI":
            wb, ws = cargar_hoja()
            ok = actualizar_fila(ws, datos["proveedor"]["ID"], datos["campo"], datos["nuevo_valor"])
            wb.save(EXCEL_FILE)
            sesiones[numero] = {"paso": "menu", "datos": {}}
            if ok:
                return f"Proveedor actualizado exitosamente!\n\n{MENU_PRINCIPAL}"
            return f"Error al actualizar. Intenta nuevamente.\n\n{MENU_PRINCIPAL}"
        elif texto.upper() == "NO":
            sesiones[numero] = {"paso": "menu", "datos": {}}
            return f"Actualizacion cancelada.\n\n{MENU_PRINCIPAL}"
        else:
            return "Responde *SI* para confirmar o *NO* para cancelar:"

    # ─ BAJA ───────────────────────────────────────────────────────────────────
    if paso == "baja_buscar":
        wb, ws = cargar_hoja()
        prov = buscar_proveedor(ws, texto)
        if not prov:
            return f"No se encontro '{texto}'. Intenta con el ID (ej: P001) o el nombre:"
        if prov["Estado"] == "Inactivo":
            sesiones[numero] = {"paso": "menu", "datos": {}}
            return f"El proveedor '{prov['Nombre']}' ya esta dado de baja.\n\n{MENU_PRINCIPAL}"
        datos["proveedor"] = prov
        sesiones[numero] = {"paso": "baja_confirmar", "datos": datos}
        return (
            f"Proveedor a dar de baja:\n\n"
            f"{formatear_proveedor(prov)}\n\n"
            f"Esta accion marcara al proveedor como *Inactivo*.\n"
            f"*SI* para confirmar / *NO* para cancelar"
        )

    if paso == "baja_confirmar":
        if texto.upper() == "SI":
            wb, ws = cargar_hoja()
            ok = actualizar_fila(ws, datos["proveedor"]["ID"], "Estado", "Inactivo")
            wb.save(EXCEL_FILE)
            sesiones[numero] = {"paso": "menu", "datos": {}}
            if ok:
                return (
                    f"Proveedor *{datos['proveedor']['Nombre']}* dado de baja.\n"
                    f"El registro queda en el sistema como Inactivo.\n\n{MENU_PRINCIPAL}"
                )
            return f"Error al dar de baja. Intenta nuevamente.\n\n{MENU_PRINCIPAL}"
        elif texto.upper() == "NO":
            sesiones[numero] = {"paso": "menu", "datos": {}}
            return f"Baja cancelada.\n\n{MENU_PRINCIPAL}"
        else:
            return "Responde *SI* para confirmar o *NO* para cancelar:"

    # ─ Fallback ───────────────────────────────────────────────────────────────
    sesiones[numero] = {"paso": "menu", "datos": {}}
    return f"No entendi. {MENU_PRINCIPAL}"


# ── Webhook Twilio ────────────────────────────────────────────────────────────

@app.route("/webhook", methods=["POST"])
def webhook():
    numero = request.form.get("From", "")
    texto  = request.form.get("Body", "").strip()

    respuesta_texto = procesar_mensaje(numero, texto)

    resp = MessagingResponse()
    resp.message(respuesta_texto)
    return str(resp)


@app.route("/", methods=["GET"])
def index():
    return "Bot de proveedores activo. Usa WhatsApp para interactuar."


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: No se encuentra '{EXCEL_FILE}'. Ejecuta 'python crear_excel.py' primero.")
    else:
        print("Servidor iniciado en http://localhost:5000")
        print("Expone con: ngrok http 5000")
        app.run(debug=True, port=5000)
