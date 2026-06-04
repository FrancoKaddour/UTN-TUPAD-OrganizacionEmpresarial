"""
Script para crear la base de datos Excel de proveedores
La Basica Pasteleria - TPI Organizacion Empresarial
Ejecutar una sola vez para generar proveedores.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

# ── Constantes de layout (todo arranca en B2) ────────────────────────────────
COL_START = 2   # columna B
ROW_TITLE = 2   # fila 2: titulo principal
ROW_SUB   = 3   # fila 3: subtitulo / fecha
ROW_HEAD  = 5   # fila 5: encabezados de columna
ROW_DATA  = 6   # fila 6 en adelante: datos

HEADERS = [
    "ID",
    "Nombre",
    "Categoria",
    "Productos",
    "Telefono",
    "Dias de Entrega",
    "Forma de Pago",
    "Estado",
    "Fecha Alta",
    "Ultima Modificacion",
]

NCOLS = len(HEADERS)
COL_END = COL_START + NCOLS - 1   # ultima columna de datos

# ── Datos ────────────────────────────────────────────────────────────────────
HOY = date.today().isoformat()

PROVEEDORES = [
    {
        "ID": "P001",
        "Nombre": "Panaderia Don Ruben",
        "Categoria": "Panaderia",
        "Productos": "Pan frances, pan lactal, pan de molde",
        "Telefono": "1154321001",
        "Dias de Entrega": "Lunes, Miercoles, Viernes",
        "Forma de Pago": "Efectivo",
        "Estado": "Activo",
        "Fecha Alta": "2024-01-10",
        "Ultima Modificacion": HOY,
    },
    {
        "ID": "P002",
        "Nombre": "Panaderia El Croissant",
        "Categoria": "Panaderia",
        "Productos": "Medialunas, vigilantes, cuernitos",
        "Telefono": "1167890002",
        "Dias de Entrega": "Martes, Jueves, Sabado",
        "Forma de Pago": "Transferencia",
        "Estado": "Activo",
        "Fecha Alta": "2024-02-15",
        "Ultima Modificacion": HOY,
    },
    {
        "ID": "P003",
        "Nombre": "Distribuidora Sabor & Arte",
        "Categoria": "Reposteria",
        "Productos": "Facturas, palmeras, canoitas, berlinesas",
        "Telefono": "1143215678",
        "Dias de Entrega": "Lunes, Jueves",
        "Forma de Pago": "Efectivo",
        "Estado": "Activo",
        "Fecha Alta": "2024-03-01",
        "Ultima Modificacion": HOY,
    },
    {
        "ID": "P004",
        "Nombre": "Coca-Cola FEMSA - Dist. Zona Sur",
        "Categoria": "Bebidas",
        "Productos": "Coca-Cola, Sprite, Fanta, agua mineral, aguas saborizadas",
        "Telefono": "0800-333-2632",
        "Dias de Entrega": "Miercoles",
        "Forma de Pago": "Credito 30 dias",
        "Estado": "Activo",
        "Fecha Alta": "2023-11-20",
        "Ultima Modificacion": HOY,
    },
    {
        "ID": "P005",
        "Nombre": "Fiambres y Delicias SA",
        "Categoria": "Fiambres y Sandwicheria",
        "Productos": "Jamon cocido, queso, lechuga, tomate, miga de pan",
        "Telefono": "1198760005",
        "Dias de Entrega": "Lunes, Miercoles, Viernes",
        "Forma de Pago": "Transferencia",
        "Estado": "Activo",
        "Fecha Alta": "2024-04-05",
        "Ultima Modificacion": HOY,
    },
    {
        "ID": "P006",
        "Nombre": "Distribuidora El Chino",
        "Categoria": "Insumos de Cocina",
        "Productos": "Harina, azucar, levadura, aceite, papel film, guantes, descartables",
        "Telefono": "1122334455",
        "Dias de Entrega": "Variable (compra presencial)",
        "Forma de Pago": "Efectivo",
        "Estado": "Activo",
        "Fecha Alta": "2023-08-01",
        "Ultima Modificacion": HOY,
    },
]

# ── Paleta de colores ─────────────────────────────────────────────────────────
C_TITULO_BG  = "1A1A2E"   # azul marino oscuro
C_TITULO_FG  = "E8C97E"   # dorado
C_SUB_BG     = "16213E"   # azul marino medio
C_SUB_FG     = "A0B4C8"   # gris azulado claro
C_HEAD_BG    = "E8C97E"   # dorado
C_HEAD_FG    = "1A1A2E"   # azul marino oscuro
C_ROW_ODD    = "FFFFFF"   # blanco
C_ROW_EVEN   = "F5F0E8"   # crema muy suave
C_ACTIVO_BG  = "D4EDDA"   # verde claro
C_ACTIVO_FG  = "155724"   # verde oscuro
C_INACT_BG   = "F8D7DA"   # rojo claro
C_INACT_FG   = "721C24"   # rojo oscuro
C_BORDER     = "C8B99A"   # borde dorado suave

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_thick():
    t = Side(style="medium", color="1A1A2E")
    return Border(left=t, right=t, top=t, bottom=t)

# ── Crear workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proveedores"
ws.sheet_view.showGridLines = False

# ── Anchos de columna ─────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 3    # margen izquierdo
WIDTHS = [8, 28, 20, 42, 16, 28, 18, 10, 14, 22]
for i, w in enumerate(WIDTHS):
    ws.column_dimensions[get_column_letter(COL_START + i)].width = w

# ── Fila 1: margen superior ───────────────────────────────────────────────────
ws.row_dimensions[1].height = 8

# ── Fila 2: TITULO ────────────────────────────────────────────────────────────
ws.row_dimensions[ROW_TITLE].height = 42
ws.merge_cells(
    start_row=ROW_TITLE, start_column=COL_START,
    end_row=ROW_TITLE,   end_column=COL_END
)
t = ws.cell(row=ROW_TITLE, column=COL_START,
            value="LA BASICA PASTELERIA  |  Gestion de Proveedores")
t.font      = Font(name="Calibri", bold=True, size=18, color=C_TITULO_FG)
t.fill      = fill(C_TITULO_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
t.border    = border_thick()

# ── Fila 3: SUBTITULO ─────────────────────────────────────────────────────────
ws.row_dimensions[ROW_SUB].height = 22
ws.merge_cells(
    start_row=ROW_SUB, start_column=COL_START,
    end_row=ROW_SUB,   end_column=COL_END
)
s = ws.cell(row=ROW_SUB, column=COL_START,
            value=f"Base de datos interna  |  Actualizado: {HOY}  |  Sistema: ChatBot WhatsApp")
s.font      = Font(name="Calibri", italic=True, size=10, color=C_SUB_FG)
s.fill      = fill(C_SUB_BG)
s.alignment = Alignment(horizontal="center", vertical="center")
s.border    = border_thick()

# ── Fila 4: espacio ───────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 6

# ── Fila 5: ENCABEZADOS ───────────────────────────────────────────────────────
ws.row_dimensions[ROW_HEAD].height = 32
for i, h in enumerate(HEADERS):
    c = ws.cell(row=ROW_HEAD, column=COL_START + i, value=h.upper())
    c.font      = Font(name="Calibri", bold=True, size=10, color=C_HEAD_FG)
    c.fill      = fill(C_HEAD_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border()

# ── Filas de datos ────────────────────────────────────────────────────────────
for r_idx, prov in enumerate(PROVEEDORES):
    row_num = ROW_DATA + r_idx
    ws.row_dimensions[row_num].height = 40
    bg = C_ROW_ODD if r_idx % 2 == 0 else C_ROW_EVEN

    for c_idx, key in enumerate(HEADERS):
        col_num = COL_START + c_idx
        valor = prov[key]
        c = ws.cell(row=row_num, column=col_num, value=valor)
        c.font      = Font(name="Calibri", size=10)
        c.fill      = fill(bg)
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if key in ("ID", "Estado", "Fecha Alta", "Ultima Modificacion") else "left")
        c.border    = border()

        # Celda Estado con color dinamico
        if key == "Estado":
            if valor == "Activo":
                c.font = Font(name="Calibri", size=10, bold=True, color=C_ACTIVO_FG)
                c.fill = fill(C_ACTIVO_BG)
            else:
                c.font = Font(name="Calibri", size=10, bold=True, color=C_INACT_FG)
                c.fill = fill(C_INACT_BG)

# ── Congelar desde la fila de datos ──────────────────────────────────────────
ws.freeze_panes = ws.cell(row=ROW_DATA, column=COL_START)

# ── Autofilter en encabezados ─────────────────────────────────────────────────
ws.auto_filter.ref = (
    f"{get_column_letter(COL_START)}{ROW_HEAD}:"
    f"{get_column_letter(COL_END)}{ROW_HEAD}"
)

# ── Guardar ───────────────────────────────────────────────────────────────────
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "proveedores.xlsx")
wb.save(OUTPUT)
print(f"[OK] '{OUTPUT}' generado con {len(PROVEEDORES)} proveedores.")
